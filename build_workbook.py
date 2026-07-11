"""Build paper_run_final.xlsx from Kalshi paper-trade CSVs. Re-runnable."""
import csv, glob, re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, DoughnutChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

PIPE = "/Users/sid/Claude/Projects/Kalshi MLB market analysis/kalshi-pipeline"
OUT = "/Users/sid/Claude/Projects/Kalshi MLB market analysis/paper_run_final.xlsx"

TEAMS = {'ATH','ATL','AZ','BAL','BOS','CHC','CIN','CLE','COL','CWS','DET','HOU',
         'KC','LAA','LAD','MIA','MIL','MIN','NYM','NYY','PHI','PIT','SD','SEA',
         'SF','STL','TB','TEX','TOR','WSH'}
NAMES = {'ATH':'Athletics','ATL':'Braves','AZ':'D-backs','BAL':'Orioles','BOS':'Red Sox',
         'CHC':'Cubs','CIN':'Reds','CLE':'Guardians','COL':'Rockies','CWS':'White Sox',
         'DET':'Tigers','HOU':'Astros','KC':'Royals','LAA':'Angels','LAD':'Dodgers',
         'MIA':'Marlins','MIL':'Brewers','MIN':'Twins','NYM':'Mets','NYY':'Yankees',
         'PHI':'Phillies','PIT':'Pirates','SD':'Padres','SEA':'Mariners','SF':'Giants',
         'STL':'Cardinals','TB':'Rays','TEX':'Rangers','TOR':'Blue Jays','WSH':'Nationals'}

def parse_ticker(t):
    m = re.match(r'KXMLBGAME-(\d{2}[A-Z]{3}\d{2})\d{4}([A-Z]+)-([A-Z]+)$', t)
    if not m: return t, '', ''
    date, teams, side = m.groups()
    for s in (2, 3):
        a, h = teams[:s], teams[s:]
        if a in TEAMS and h in TEAMS:
            return f"{NAMES.get(a,a)} @ {NAMES.get(h,h)}", NAMES.get(side, side), date.title()
    return t, NAMES.get(side, side), date.title()

rows = []
seen = set()
for f in sorted(glob.glob(f"{PIPE}/paper_trades*.csv")):
    with open(f) as fh:
        for r in csv.DictReader(fh):
            key = (r['signal_time'], r['ticker'], r['entry_price'], r['exit_price'])
            if key in seen or not r.get('exit_price'): continue
            seen.add(key)
            rows.append(r)
rows.sort(key=lambda r: r['signal_time'])

ARIAL = 'Arial'
NAVY, LNAVY = '1F3864', 'D9E2F3'
GREEN_F, GREEN_BG = '006100', 'C6EFCE'
RED_F, RED_BG = '9C0006', 'FFC7CE'
GRAY = 'F2F2F2'
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(c, bold=False, size=10, color='000000', bg=None, align='left', wrap=False, italic=False):
    c.font = Font(name=ARIAL, bold=bold, size=size, color=color, italic=italic)
    if bg: c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)

wb = Workbook()

# ============================ Trade Log ============================
log = wb.active
log.title = 'Trade Log'
headers = ['#', 'Date', 'Time (ET)', 'Game', 'Team Bought', 'Trigger Play',
           'Entry (¢)', 'Exit (¢)', 'Exit Type', 'Fees ($)', 'P&L (¢)', 'P&L ($)',
           'Result', 'Cumulative P&L ($)']
for j, h in enumerate(headers, 1):
    c = log.cell(1, j, h)
    style(c, bold=True, color='FFFFFF', bg=NAVY, align='center', wrap=True)
    c.border = BORDER
log.freeze_panes = 'A2'
log.row_dimensions[1].height = 30

first, last = 2, 1 + len(rows)
for i, r in enumerate(rows):
    row = first + i
    game, team, date = parse_ticker(r['ticker'])
    dt = datetime.fromisoformat(r['signal_time'].replace('Z', '+00:00'))
    et = dt - timedelta(hours=4)
    ev = (r.get('event') or '').replace('_', ' ').title() or 'Batting event'
    vals = [i + 1, et.strftime('%b %d'), et.strftime('%I:%M %p'), game, team, ev,
            int(r['entry_price']), int(r['exit_price']), r['exit_mode'],
            float(r['fees']), float(r['pnl_cents_conservative'])]
    for j, v in enumerate(vals, 1):
        c = log.cell(row, j, v)
        style(c, align='center' if j not in (4,) else 'left',
              bg=GRAY if i % 2 else None)
        c.border = BORDER
    log.cell(row, 10).number_format = '$#,##0.00'
    log.cell(row, 11).number_format = '#,##0;(#,##0);-'
    c = log.cell(row, 12, f'=K{row}/100')
    c.number_format = '$#,##0.00;($#,##0.00);-'
    style(c, align='center', bg=GRAY if i % 2 else None); c.border = BORDER
    c = log.cell(row, 13, f'=IF(K{row}>0,"WIN",IF(K{row}<0,"LOSS","FLAT"))')
    style(c, bold=True, align='center', bg=GRAY if i % 2 else None); c.border = BORDER
    c = log.cell(row, 14, f'=SUM($L${first}:L{row})')
    c.number_format = '$#,##0.00;($#,##0.00);-'
    style(c, align='center', bg=GRAY if i % 2 else None); c.border = BORDER

for rng, ops in [(f'K{first}:L{last}', None), (f'M{first}:M{last}', 'wl'), (f'N{first}:N{last}', None)]:
    log.conditional_formatting.add(rng, CellIsRule(
        operator='greaterThan', formula=['0'] if ops is None else ['"zz"'],
        font=Font(name=ARIAL, color=GREEN_F, bold=True),
        fill=PatternFill('solid', start_color=GREEN_BG)) if ops is None else
        CellIsRule(operator='equal', formula=['"WIN"'],
        font=Font(name=ARIAL, color=GREEN_F, bold=True),
        fill=PatternFill('solid', start_color=GREEN_BG)))
    log.conditional_formatting.add(rng, CellIsRule(
        operator='lessThan', formula=['0'],
        font=Font(name=ARIAL, color=RED_F, bold=True),
        fill=PatternFill('solid', start_color=RED_BG)) if ops is None else
        CellIsRule(operator='equal', formula=['"LOSS"'],
        font=Font(name=ARIAL, color=RED_F, bold=True),
        fill=PatternFill('solid', start_color=RED_BG)))

widths = [5, 9, 10, 24, 14, 14, 9, 9, 10, 9, 9, 10, 9, 16]
for j, w in enumerate(widths, 1):
    log.column_dimensions[get_column_letter(j)].width = w

# helper columns for win/loss overlay bars (hidden)
log.cell(1, 16, 'Win ¢'); log.cell(1, 17, 'Loss ¢')
for i in range(len(rows)):
    row = first + i
    log.cell(row, 16, f'=IF(K{row}>0,K{row},"")')
    log.cell(row, 17, f'=IF(K{row}<0,K{row},"")')
log.column_dimensions['P'].hidden = True
log.column_dimensions['Q'].hidden = True

# ============================ Dashboard ============================
db = wb.create_sheet('Dashboard', 0)
db.sheet_view.showGridLines = False
db.merge_cells('B2:M2')
c = db['B2']; c.value = 'KALSHI MLB PAPER-TRADING RESULTS'
style(c, bold=True, size=18, color='FFFFFF', bg=NAVY, align='center')
for col in 'CDEFGHIJKLM':
    style(db[f'{col}2'], bg=NAVY)
db.row_dimensions[2].height = 34
db.merge_cells('B3:M3')
c = db['B3']
c.value = ('Simulated (no real money) live trading of MLB moneyline markets on Kalshi. '
           'The bot buys a team the moment it gets a hit, walk, or home run — betting the '
           'market is slow to react — holds 5 minutes, then sells. See "How It Works" tab.')
style(c, italic=True, size=10, color='404040', align='center', wrap=True)
db.row_dimensions[3].height = 30

TL = "'Trade Log'"
kpis = [
    ('Total P&L', f'=SUM({TL}!L{first}:L{last})', '$#,##0.00;($#,##0.00)', 'money'),
    ('Trades', f'=COUNT({TL}!A{first}:A{last})', '0', None),
    ('Win Rate', f'=COUNTIF({TL}!M{first}:M{last},"WIN")/COUNT({TL}!A{first}:A{last})', '0.0%', None),
    ('Avg Win', f'=IFERROR(AVERAGEIF({TL}!L{first}:L{last},">0"),0)', '$#,##0.00', 'good'),
    ('Avg Loss', f'=IFERROR(AVERAGEIF({TL}!L{first}:L{last},"<0"),0)', '$#,##0.00;($#,##0.00)', 'bad'),
    ('Profit Factor', f'=IFERROR(SUMIF({TL}!L{first}:L{last},">0")/ABS(SUMIF({TL}!L{first}:L{last},"<0")),0)', '0.00x', None),
    ('Total Fees', f'=SUM({TL}!J{first}:J{last})', '$#,##0.00', None),
    ('Best Trade', f'=MAX({TL}!L{first}:L{last})', '$#,##0.00', 'good'),
    ('Worst Trade', f'=MIN({TL}!L{first}:L{last})', '$#,##0.00;($#,##0.00)', 'bad'),
]
kpi_cols = ['B','C','D','E','F','G','H','I','J']
for (label, formula, fmt, tone), col in zip(kpis, kpi_cols):
    c = db[f'{col}5']; c.value = label
    style(c, bold=True, size=9, color='FFFFFF', bg='44546A', align='center', wrap=True)
    c.border = BORDER
    c = db[f'{col}6']; c.value = formula; c.number_format = fmt
    tone_color = {'good': GREEN_F, 'bad': RED_F}.get(tone, '000000')
    style(c, bold=True, size=13, color=tone_color, bg=LNAVY, align='center')
    c.border = BORDER
    db.column_dimensions[col].width = 13
db.row_dimensions[5].height = 24
db.row_dimensions[6].height = 30

c = db['B8']; c.value = 'Quick read:'
style(c, bold=True, size=11)
db.merge_cells('C8:M8')
c = db['C8']
c.value = ('=IF(SUM(' + TL + f'!L{first}:L{last})>=0,'
           '"The run is PROFITABLE so far — but the sample is still small; judge after 100+ trades.",'
           '"The run is DOWN so far — small samples swing hard; the backtest edge (~+1 to +3¢/trade) '
           'needs 100+ trades to show through noise.")')
style(c, italic=True, size=10, color='404040', wrap=True)

# exit-mode + W/L helper table (feeds doughnut)
db['O5'] = 'Wins'; db['P5'] = f'=COUNTIF({TL}!M{first}:M{last},"WIN")'
db['O6'] = 'Losses'; db['P6'] = f'=COUNTIF({TL}!M{first}:M{last},"LOSS")'
for cc in ('O5','P5','O6','P6'): style(db[cc], size=9)

# charts
line = LineChart()
line.title = 'Cumulative P&L ($) — running total after each trade'
line.style = 12; line.height = 8; line.width = 16
line.y_axis.numFmt = '$#,##0.00'
line.x_axis.title = 'Trade #'
data = Reference(log, min_col=14, min_row=1, max_row=last)
cats = Reference(log, min_col=1, min_row=first, max_row=last)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.legend = None
db.add_chart(line, 'B10')

bars = BarChart()
bars.type = 'col'; bars.overlap = 100; bars.gapWidth = 60
bars.title = 'P&L per trade (¢) — green wins, red losses'
bars.height = 8; bars.width = 16
bars.x_axis.title = 'Trade #'
win = Series(Reference(log, min_col=16, min_row=first, max_row=last), title='Win')
loss = Series(Reference(log, min_col=17, min_row=first, max_row=last), title='Loss')
win.graphicalProperties.solidFill = '70AD47'
loss.graphicalProperties.solidFill = 'C00000'
bars.append(win); bars.append(loss)
bars.set_categories(cats)
bars.legend = None
db.add_chart(bars, 'B27')

dough = DoughnutChart()
dough.title = 'Win / Loss split'
dough.height = 8; dough.width = 8
d = Reference(db, min_col=16, min_row=5, max_row=6)
labels = Reference(db, min_col=15, min_row=5, max_row=6)
dough.add_data(d); dough.set_categories(labels)
from openpyxl.chart.series import DataPoint
pts = [DataPoint(idx=0), DataPoint(idx=1)]
pts[0].graphicalProperties.solidFill = '70AD47'
pts[1].graphicalProperties.solidFill = 'C00000'
dough.series[0].data_points = pts
db.add_chart(dough, 'K10')

# ============================ How It Works ============================
hw = wb.create_sheet('How It Works')
hw.sheet_view.showGridLines = False
hw.column_dimensions['B'].width = 26
hw.column_dimensions['C'].width = 95
content = [
    ('title', 'HOW THIS PAPER-TRADING RUN WORKS', ''),
    ('h', 'The strategy in one sentence', ''),
    ('p', '', 'When a batter gets a hit, walk, or home run, his team\'s chance of winning goes up — '
     'but Kalshi\'s prediction-market prices take a few minutes to fully adjust. The bot buys the '
     'batting team\'s "will they win?" contract seconds after the play, waits 5 minutes for the '
     'market to catch up, then sells. The hoped-for profit is that few-cent catch-up move.'),
    ('h', 'Why "paper" trading?', ''),
    ('p', '', 'No real money is used. The bot watches real live market data and simulates what would '
     'have happened to real orders — including realistic queue waiting, fees, and having to accept '
     'a worse price when an order doesn\'t fill in time. It\'s a dress rehearsal to measure whether '
     'the strategy actually earns money before risking anything.'),
    ('h', 'Where the numbers come from', ''),
    ('p', '', 'A research backtest over the full 2026 season (882 games, ~66,000 market reactions) '
     'suggested an edge of roughly +1 to +3 cents per trade after spread costs, strongest when '
     'holding 5-10 minutes. Each live trade here risks 10 contracts (~$5-9 at stake per trade).'),
    ('h', 'Column definitions (Trade Log)', ''),
    ('d', 'Trigger Play', 'The batting event (single, walk, home run...) that fired the buy signal.'),
    ('d', 'Entry / Exit (¢)', 'Contract prices in cents. A contract pays 100¢ if the team wins, 0¢ if not. '
     'Buying at 47¢ means the market thought the team had ~47% chance to win.'),
    ('d', 'Exit Type', '"maker" = sold patiently at our asking price (no fee). '
     '"taker" = order timed out after 2 min, so the bot crossed the spread and paid a small fee to get out.'),
    ('d', 'P&L', '(Exit − Entry) × 10 contracts − fees. 1¢ of price move = 10¢ of P&L.'),
    ('d', 'Result', 'WIN if the trade made money, LOSS if it lost.'),
    ('d', 'Cumulative P&L', 'Running total of all trades so far — the line chart on the Dashboard.'),
    ('h', 'What to look for', ''),
    ('p', '', '1) Win rate near or above 50%.  2) Average win at least as big as average loss '
     '(Profit Factor above 1.00x means gross wins exceed gross losses).  3) A cumulative P&L line '
     'that trends up over MANY trades. A handful of trades means nothing either way — variance '
     'dominates until roughly 100+ trades.'),
    ('h', 'Honest caveats', ''),
    ('p', '', 'The simulated fills are estimates (real orders might fill less often or at worse '
     'prices). Buying only after good offensive events also means buying into momentum — several '
     'early losses came from the other team immediately answering back. The whole point of this '
     'run is to find out, cheaply, whether the backtest edge survives contact with live markets.'),
]
r = 2
for kind, b, ctext in content:
    if kind == 'title':
        hw.merge_cells(f'B{r}:C{r}')
        c = hw[f'B{r}']; c.value = b
        style(c, bold=True, size=15, color='FFFFFF', bg=NAVY, align='center')
        hw.row_dimensions[r].height = 28
        r += 2
    elif kind == 'h':
        c = hw[f'B{r}']; c.value = b
        hw.merge_cells(f'B{r}:C{r}')
        style(c, bold=True, size=12, color=NAVY)
        r += 1
    elif kind == 'p':
        hw.merge_cells(f'B{r}:C{r}')
        c = hw[f'B{r}']; c.value = ctext
        style(c, size=10, wrap=True)
        hw.row_dimensions[r].height = 14 * (len(ctext) // 110 + 2)
        r += 2
    elif kind == 'd':
        c = hw[f'B{r}']; c.value = b
        style(c, bold=True, size=10, color=NAVY, align='right')
        c = hw[f'C{r}']; c.value = ctext
        style(c, size=10, wrap=True)
        hw.row_dimensions[r].height = 14 * (len(ctext) // 95 + 1) + 4
        r += 1

wb.save(OUT)
print(f"saved {OUT} with {len(rows)} trades")
